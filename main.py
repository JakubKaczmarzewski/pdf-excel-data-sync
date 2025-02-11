import re
import openpyxl
import pypdf
import os
from urllib.parse import unquote
from pathlib import Path
from openpyxl.reader.excel import load_workbook
import tabula


def create_full_path(path_prefix: str, filename: str) -> str:
    """Creating and returning full path to the file. Joining `path_prefix` with `filename`"""

    filename = unquote(filename)
    full_path = os.path.join(path_prefix, filename)
    return full_path


def move_file(file_name: str, path_to_move: str) -> None:
    """Move file to different location."""

    pass


def read_pdf(pdf_path: str) -> str:
    """
    Read data from a PDF file and store it as a single string.
    :param pdf_path: Path to the PDF file.
    :return: Extracted text from the PDF as a string.
    """

    # print(pdf_path)
    pdf_as_str = ""

    try:
        with open(pdf_path, 'rb') as pdf:
            pdf_reader = pypdf.PdfReader(pdf)  # Create a PDF reader object

            # Iterate through each page in the PDF
            for page_num in range(len(pdf_reader.pages)):
                page = pdf_reader.pages[page_num]

                # Extract text from the page
                text = page.extract_text()
                if text:
                    pdf_as_str += text

    except FileNotFoundError:
        print(f"Error: The file at {pdf_path} was not found.")
        return ""
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

    return pdf_as_str


def extract_document_code(text: str, search_pattern: list = ("Document No.:", "Internal Ref. Nr.:")) -> str:
    """
    Extract the document code that is preceded by a user-defined pattern
    and ends with a newline character.

    :param text: The text in which to search for the document code.
    :param search_pattern: The list or regex pattern to search for before the document code
            (default is "Document No.:", "Internal Ref. Nr.:").
    :return: The extracted document code or None if not found.
    """

    for item in search_pattern:
        search_regex = re.escape(item) + r"([^\n]+)"

        match = re.search(search_regex, text)

        if match:
            return match.group(1).strip()

    return None


def extract_product_codes(pdf_path: str, column_candidates: list = ['12NC', '10NC', 'Product Code', 'Model']) -> list:
    """
    Look for tables in a PDF file and extract product codes from the specified column.

    Using the Tabula module, search for tables in a PDF file and extract product codes
    from the first matching column name in the provided list of candidates.

    :param pdf_path: Total path to the PDF file to process.
    :param column_candidates: List of column names to search for in order of priority.
    :return: List of extracted product codes.
    """
    try:
        # Read all tables from the PDF
        dfs = tabula.read_pdf(pdf_path, pages='all', multiple_tables=True)

        extracted_codes = []

        # Iterate over all detected tables
        for df in dfs:
            # Check if the dataframe is valid
            if df is not None and not df.empty:
                # Find the first matching column
                for column_name in column_candidates:
                    if column_name in df.columns:
                        # Extend the list with the values from the matching column
                        extracted_codes.extend(df[column_name].dropna().astype(str).tolist())
                        break  # Stop looking for other columns in this table

        return extracted_codes

    except Exception as e:
        print(f"An error occurred while processing the PDF: {e}")
        return []


def process_excel_and_pdfs(excel_file_path: str, pdf_link_prefix: str = "", start_row: int = 2, end_row=None) -> list:
    """
    Process the Excel file and for each row, extract data from the PDF linked in the Excel.
    Write the extracted document code into the next column of the same row (column need to exist and have header).
    :param excel_file_path: Path to the Excel file.
    :param pdf_link_prefix: The prefix to be added to PDF links if they are partial.
    :param start_row: Number of the row where program starts to process, value 2 by default.
    :param end_row: Number of the row where program ends, if not specified by user program process whole file.
    :return: Lists with numbers and file_names of unprocessed rows.
    """
    print(pdf_link_prefix)
    wb_obj = openpyxl.load_workbook(excel_file_path)
    active_sheet = wb_obj.active

    logs = list()

    if end_row is None:
        end_row = active_sheet.max_row

    # Iterate over the rows in the Excel file
    for row in active_sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=active_sheet.max_column):
        product_code = row[0].value  # Assuming the product code is in column A
        pdf_link = row[1].hyperlink.target if row[1].hyperlink else None  # Extract the hyperlink from column B
        if pdf_link:  # Only process if there's a valid hyperlink
            print(f"Processing PDF for Product Code: {product_code} - {pdf_link}")

            # Decode the URL-encoded characters in the file path
            decoded_pdf_link = unquote(pdf_link)

            # Join the prefix and the decoded path
            full_path = os.path.join(pdf_link_prefix, decoded_pdf_link.lstrip("\\/"))

            print(f"Decoded and full path: {full_path}")

            pdf_text = read_pdf(full_path)

            # Extract the document code from the PDF text
            document_code = extract_document_code(pdf_text)

            if document_code:
                print(f"Extracted Document Code: {document_code}")

                # Write the document code to the next column (C)
                row[2].value = document_code
            else:
                unextracted_doc = (int(row[0].row), decoded_pdf_link)
                logs.append(unextracted_doc)

    # Save the updated Excel file after processing
    wb_obj.save(f"updated_{os.path.basename(excel_file_path)}")
    return logs


def process_folder_write_pdfs_data_to_excel(file_path: str, process_excel_path: str, excel_save_path: str = None):
    """Open folder read all pdf files in it, for each one extract data and write it to Excel file, after that move file.


    """
    #  Load Excel file to write data
    process_excel_path = unquote(process_excel_path)
    excel_to_write = openpyxl.load_workbook(process_excel_path)
    active_sheet = excel_to_write.active

    # Processing folder
    num_of_pdfs = 0
    for filename in os.listdir(file_path):
        if filename.endswith(".pdf"):
            full_file_path = create_full_path(file_path, filename)
            if os.path.isfile(full_file_path):
                print(f"Processing file: {full_file_path}")
                num_of_pdfs += 1
                pdf_as_str = read_pdf(full_file_path)
                extracted_doc_code = extract_document_code(pdf_as_str)  # Extracting document code form PDF file

                # Extracting product codes
                product_codes = extract_product_codes(full_file_path)

                for item in product_codes:
                    print(item)

                # Generating link to filename

                # Write all items to excel
                last_row = active_sheet.max_row + 1  # Finding last Excel row for each file.
                print(f"Last row is {last_row}")
                for index, code in enumerate(product_codes, start=last_row):
                    active_sheet.cell(row=index, column=1, value=code)  # Column 1: 12NC code
                    active_sheet.cell(row=index, column=2, value=full_file_path)  # Column 2: Full path to file
                    active_sheet.cell(row=index, column=3, value=extracted_doc_code)  # Column 3: Document code

                # Move processed file to different location
                print("========NEXT ITEM=======")

    # Save the updated Excel file after processing
    if excel_save_path is None:
        excel_save_path = process_excel_path
        print(f"Saving file at the location: {excel_save_path}")
    excel_to_write.save(os.path.basename(excel_save_path))
    print(f"Number of processed pdf files: {num_of_pdfs}")


if __name__ == '__main__':
    # Main task

    # pdf_link_prefix = r""  # Here type PDF prefix file if needed
    # excel_file = r""  # Here type path to the Excel file (Necessary)
    # process_excel_and_pdfs(excel_file, pdf_link_prefix)

    # process_folder_write_pdfs_data_to_excel()  # New feature
    pass
