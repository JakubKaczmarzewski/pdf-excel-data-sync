import os
import re
import tabula
import pandas as pd
from openpyxl import load_workbook
from urllib.parse import unquote
import pypdf


def read_pdf(pdf_path: str) -> str:
    """Read data from a PDF file and return it as a single string."""
    pdf_as_str = ""

    try:
        with open(pdf_path, 'rb') as pdf:
            pdf_reader = pypdf.PdfReader(pdf)

            for page in pdf_reader.pages:
                text = page.extract_text()
                if text:
                    pdf_as_str += text

    except FileNotFoundError:
        print(f"Error: The file at {pdf_path} was not found.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

    return pdf_as_str


def extract_document_code(text: str, search_pattern=("Document No.:", "Internal Ref. Nr.:")) -> str:
    """Extract the document code based on patterns from the PDF text."""
    for item in search_pattern:
        search_regex = re.escape(item) + r"([^\n]+)"
        match = re.search(search_regex, text)

        if match:
            return match.group(1).strip()

    return None


def link_generator(path_prefix: str, filename: str) -> str:
    """Generate a link by combining a prefix and a filename."""
    return os.path.join(path_prefix, filename)


def process_pdf_and_write_to_excel(pdf_folder: str, excel_file: str, output_excel_file: str, pdf_link_prefix: str):
    """
    Process each PDF file in a folder, extract data, and write it into an Excel file.

    :param pdf_folder: Folder containing PDF files.
    :param excel_file: Path to the input Excel file.
    :param output_excel_file: Path to save the updated Excel file.
    :param pdf_link_prefix: Prefix for generating links to PDF files.
    """
    # Load the Excel workbook
    workbook = load_workbook(excel_file)
    sheet = workbook.active

    # Iterate through PDF files in the specified folder
    for filename in os.listdir(pdf_folder):
        if filename.endswith(".pdf"):
            pdf_path = os.path.join(pdf_folder, filename)

            # Read PDF and extract text
            pdf_text = read_pdf(pdf_path)

            # Extract document code
            document_code = extract_document_code(pdf_text)

            # Extract 12NC codes from the PDF table
            dfs = tabula.read_pdf(pdf_path, pages='all', multiple_tables=True)
            if dfs and '12NC' in dfs[0].columns:
                new_data = dfs[0]['12NC']
            else:
                new_data = []

            # Find the next available row in the Excel sheet
            last_row = sheet.max_row + 1

            # Write 12NC codes and document code into Excel
            for idx, code in enumerate(new_data, start=last_row):
                sheet.cell(row=idx, column=1, value=code)  # Column 1: 12NC code
                sheet.cell(row=idx, column=3, value=document_code)  # Column 3: Document code

            # Generate and write link to PDF
            link = link_generator(pdf_link_prefix, filename)
            for idx in range(last_row, last_row + len(new_data)):
                sheet.cell(row=idx, column=2, value=link)  # Column 2: Link

    # Save the updated Excel file
    workbook.save(output_excel_file)
    print(f"Updated Excel file saved to {output_excel_file}")


if __name__ == '__main__':
    # Define paths and parameters
    pass
