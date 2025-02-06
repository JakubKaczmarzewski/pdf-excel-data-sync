import tabula
import pandas as pd
import openpyxl
from openpyxl import writer
from openpyxl.reader.excel import load_workbook
import os


def link_generator(path_prefix: str, filename: str):
    pass


# Ścieżki do pliku PDF i pliku Excel
pdf_path = r"C:\Users\670336256\Desktop\2021A0117.pdf"
excel_file = r"C:\Users\670336256\Desktop\ExcelTestFile.xlsx"
updated_excel_file = r"C:\Users\670336256\Desktop\UpdatedExcelFile.xlsx"

# Wczytanie tabel z pliku PDF
dfs = tabula.read_pdf(pdf_path, pages='all', multiple_tables=True)
print(dfs)

# Wydzielenie interesującej kolumny z pierwszej tabeli w PDF
new_data = dfs[0]['12NC']
print("New Data (from PDF):")
print(new_data)

# Wczytanie istniejącego pliku Excel
workbook = load_workbook(excel_file)
sheet = workbook.active  # Wybór aktywnego arkusza

# Znalezienie ostatniego wiersza w pierwszej kolumnie
last_row = sheet.max_row

# Dodanie wartości z kolumny '12NC' poniżej istniejących danych
for idx, value in enumerate(new_data, start=last_row + 1):  # Zaczynamy od pierwszego pustego wiersza
    sheet.cell(row=idx, column=1, value=value)  # Wstawienie wartości w pierwszej kolumnie

# Zapisanie zmodyfikowanego pliku Excel
workbook.save(updated_excel_file)

print(f"Updated Excel file saved to {updated_excel_file}")


